"""Relatórios do MVP."""

from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, Optional
import sqlite3

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from rules import classify_rubric


def _where_company(company_id: Optional[int]) -> tuple[str, list[Any]]:
    if company_id:
        return "WHERE r.company_id = ?", [company_id]
    return "", []


def load_analytic(conn: sqlite3.Connection, rules: Dict[str, Any], company_id: Optional[int] = None) -> pd.DataFrame:
    where, params = _where_company(company_id)
    sql = f"""
        SELECT
            e.tp_insc,
            e.nr_insc,
            COALESCE(e.cnpj_completo, e.nr_insc) AS cnpj_completo,
            COALESCE(e.razao_social, 'Empresa CNPJ base ' || e.nr_insc) AS razao_social,
            r.company_id,
            r.event_type,
            r.per_apur,
            r.ano,
            r.mes,
            r.cpf_trab,
            r.matricula,
            r.cod_categ,
            r.ide_dm_dev,
            r.cod_rubr,
            r.ide_tab_rubr,
            rb.dsc_rubr,
            rb.nat_rubr,
            rb.tp_rubr,
            rb.cod_inc_cp,
            rb.cod_inc_irrf,
            rb.cod_inc_fgts,
            r.qtd_rubr,
            r.fator_rubr,
            r.vr_rubr,
            r.tp_insc_estab,
            r.nr_insc_estab,
            r.cod_lotacao,
            r.source_file
        FROM remuneracoes r
        JOIN empresas e ON e.id = r.company_id
        LEFT JOIN rubricas rb
          ON rb.company_id = r.company_id
         AND rb.cod_rubr = r.cod_rubr
         AND rb.ide_tab_rubr = r.ide_tab_rubr
        {where}
        ORDER BY r.ano, r.mes, r.cod_rubr, r.cpf_trab
    """
    df = pd.read_sql_query(sql, conn, params=params)
    if df.empty:
        return df

    classifications = df.apply(lambda row: classify_rubric(row.to_dict(), rules), axis=1, result_type="expand")
    df = pd.concat([df, classifications], axis=1)
    return df


def apply_filters(df: pd.DataFrame, start_period: Optional[str] = None, end_period: Optional[str] = None, statuses: Optional[list[str]] = None) -> pd.DataFrame:
    if df.empty:
        return df
    out = df.copy()
    if start_period:
        out = out[out["per_apur"].fillna("") >= start_period]
    if end_period:
        out = out[out["per_apur"].fillna("") <= end_period]
    if statuses:
        out = out[out["status_analise"].isin(statuses)]
    return out


def add_credit_estimate(df: pd.DataFrame, estimated_rate_percent: float) -> pd.DataFrame:
    if df.empty:
        return df
    out = df.copy()
    out["valor_base_questionado"] = out.apply(
        lambda row: float(row["vr_rubr"] or 0.0)
        if row.get("status_analise") in {"POTENCIAL_CREDITO_ANALISAR", "REVISAR_NOMENCLATURA_BASE_CP"}
        else 0.0,
        axis=1,
    )
    out["credito_estimado"] = out["valor_base_questionado"] * (float(estimated_rate_percent) / 100.0)
    return out


def summarize_by_month_rubric(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()
    group_cols = [
        "tp_insc",
        "nr_insc",
        "cnpj_completo",
        "razao_social",
        "company_id",
        "ano",
        "mes",
        "per_apur",
        "cod_rubr",
        "ide_tab_rubr",
        "dsc_rubr",
        "nat_rubr",
        "cod_inc_cp",
        "status_analise",
    ]
    return (
        df.groupby(group_cols, dropna=False)
        .agg(
            valor_pago=("vr_rubr", "sum"),
            valor_base_questionado=("valor_base_questionado", "sum"),
            credito_estimado=("credito_estimado", "sum"),
            trabalhadores=("cpf_trab", pd.Series.nunique),
            lancamentos=("vr_rubr", "count"),
        )
        .reset_index()
        .sort_values(["ano", "mes", "cod_rubr"], na_position="last")
    )


def summarize_credit_by_month(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()

    out = df.copy()
    out["rubrica_com_credito"] = out.apply(
        lambda row: row.get("cod_rubr") if float(row.get("valor_base_questionado") or 0.0) > 0 else None,
        axis=1,
    )
    group_cols = [
        "tp_insc",
        "nr_insc",
        "cnpj_completo",
        "razao_social",
        "company_id",
        "ano",
        "mes",
        "per_apur",
    ]
    monthly = (
        out.groupby(group_cols, dropna=False)
        .agg(
            valor_total_folha=("vr_rubr", "sum"),
            valor_base_questionada=("valor_base_questionado", "sum"),
            credito_estimado=("credito_estimado", "sum"),
            rubricas=("cod_rubr", pd.Series.nunique),
            rubricas_com_credito=("rubrica_com_credito", pd.Series.nunique),
            trabalhadores=("cpf_trab", pd.Series.nunique),
            lancamentos=("vr_rubr", "count"),
        )
        .reset_index()
    )
    monthly = monthly.rename(columns={"per_apur": "periodo_apurado"})
    monthly["mes"] = monthly.apply(
        lambda row: "13"
        if (pd.isna(row.get("mes")) or str(row.get("mes")).strip() == "") and len(str(row.get("periodo_apurado") or "")) == 4
        else row.get("mes"),
        axis=1,
    )
    monthly["_mes_ordem"] = pd.to_numeric(monthly["mes"], errors="coerce").fillna(99)
    monthly["status_credito"] = monthly["credito_estimado"].apply(
        lambda value: "COM_CREDITO_ESTIMADO" if float(value or 0.0) > 0 else "SEM_CREDITO_ESTIMADO"
    )
    monthly = monthly.sort_values(["ano", "_mes_ordem", "periodo_apurado"], na_position="last").drop(columns=["_mes_ordem"])
    return monthly[
        [
            "tp_insc",
            "nr_insc",
            "cnpj_completo",
            "razao_social",
            "company_id",
            "ano",
            "mes",
            "periodo_apurado",
            "status_credito",
            "valor_total_folha",
            "valor_base_questionada",
            "credito_estimado",
            "rubricas",
            "rubricas_com_credito",
            "trabalhadores",
            "lancamentos",
        ]
    ]


def load_payment_summary(conn: sqlite3.Connection, company_id: Optional[int] = None) -> pd.DataFrame:
    params: list[Any] = []
    where = ""
    if company_id:
        where = "WHERE p.company_id = ?"
        params.append(company_id)
    return pd.read_sql_query(
        f"""
        SELECT
            p.company_id,
            p.per_apur AS periodo_apurado,
            GROUP_CONCAT(DISTINCT SUBSTR(p.dt_pgto, 1, 7)) AS periodos_pagamento,
            MIN(p.dt_pgto) AS primeira_data_pagamento,
            MAX(p.dt_pgto) AS ultima_data_pagamento,
            SUM(COALESCE(p.vr_liq, 0)) AS valor_liquido_pago_s1210,
            COUNT(*) AS pagamentos_s1210
        FROM pagamentos p
        {where}
        GROUP BY p.company_id, p.per_apur
        ORDER BY p.per_apur
        """,
        conn,
        params=params,
    )


def summarize_by_rubric(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()
    group_cols = [
        "tp_insc",
        "nr_insc",
        "cnpj_completo",
        "razao_social",
        "cod_rubr",
        "ide_tab_rubr",
        "dsc_rubr",
        "nat_rubr",
        "cod_inc_cp",
        "status_analise",
        "motivo_analise",
    ]
    return (
        df.groupby(group_cols, dropna=False)
        .agg(
            valor_pago=("vr_rubr", "sum"),
            valor_base_questionado=("valor_base_questionado", "sum"),
            credito_estimado=("credito_estimado", "sum"),
            competencias=("per_apur", pd.Series.nunique),
            trabalhadores=("cpf_trab", pd.Series.nunique),
            lancamentos=("vr_rubr", "count"),
        )
        .reset_index()
        .sort_values(["valor_pago", "cod_rubr"], ascending=[False, True], na_position="last")
    )


def build_inss_base_map(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()

    out = summarize_by_rubric(df).copy()

    def base_status(row: pd.Series) -> str:
        if row.get("status_analise") == "SEM_TABELA_S1010":
            return "PENDENTE_S1010"
        cod_inc_cp = str(row.get("cod_inc_cp") or "").strip()
        if not cod_inc_cp or cod_inc_cp in {"00", "01"} or cod_inc_cp.startswith("9"):
            return "NAO_COMPOE_BASE_CP"
        return "COMPOE_BASE_CP"

    def credit_origin(row: pd.Series) -> str:
        if row.get("status_analise") == "POTENCIAL_CREDITO_ANALISAR":
            return "Rubrica candidata com codIncCP indicando cobrança previdenciária."
        if row.get("status_analise") == "SEM_TABELA_S1010":
            return "Sem S-1010 para validar natureza e incidência."
        return str(row.get("motivo_analise") or "")

    out["situacao_base_cp"] = out.apply(base_status, axis=1)
    out["origem_credito"] = out.apply(credit_origin, axis=1)
    out["valor_em_base_cp"] = out.apply(
        lambda row: row["valor_pago"] if row["situacao_base_cp"] == "COMPOE_BASE_CP" else 0.0,
        axis=1,
    )
    out["valor_sem_s1010"] = out.apply(
        lambda row: row["valor_pago"] if row["situacao_base_cp"] == "PENDENTE_S1010" else 0.0,
        axis=1,
    )
    return out[
        [
            "cod_rubr",
            "ide_tab_rubr",
            "dsc_rubr",
            "nat_rubr",
            "cod_inc_cp",
            "situacao_base_cp",
            "status_analise",
            "valor_pago",
            "valor_em_base_cp",
            "valor_base_questionado",
            "credito_estimado",
            "valor_sem_s1010",
            "trabalhadores",
            "lancamentos",
            "origem_credito",
        ]
    ]


def summarize_by_year(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()
    group_cols = ["tp_insc", "nr_insc", "cnpj_completo", "razao_social", "ano", "status_analise"]
    return (
        df.groupby(group_cols, dropna=False)
        .agg(
            valor_pago=("vr_rubr", "sum"),
            valor_base_questionado=("valor_base_questionado", "sum"),
            credito_estimado=("credito_estimado", "sum"),
            rubricas=("cod_rubr", pd.Series.nunique),
            trabalhadores=("cpf_trab", pd.Series.nunique),
            lancamentos=("vr_rubr", "count"),
        )
        .reset_index()
        .sort_values(["ano", "status_analise"], na_position="last")
    )


def load_rubrics(conn: sqlite3.Connection, company_id: Optional[int] = None) -> pd.DataFrame:
    params: list[Any] = []
    where = ""
    if company_id:
        where = "WHERE rb.company_id = ?"
        params.append(company_id)
    return pd.read_sql_query(
        f"""
        SELECT
            e.tp_insc,
            e.nr_insc,
            COALESCE(e.cnpj_completo, e.nr_insc) AS cnpj_completo,
            COALESCE(e.razao_social, 'Empresa CNPJ base ' || e.nr_insc) AS razao_social,
            rb.*
        FROM rubricas rb
        JOIN empresas e ON e.id = rb.company_id
        {where}
        ORDER BY rb.nat_rubr, rb.cod_rubr
        """,
        conn,
        params=params,
    )


def load_imports(conn: sqlite3.Connection) -> pd.DataFrame:
    return pd.read_sql_query(
        """
        SELECT ai.imported_at, ai.filename, ai.event_type, e.nr_insc, e.cnpj_completo,
               COALESCE(e.razao_social, 'Empresa CNPJ base ' || e.nr_insc) AS razao_social,
               ai.qtd_rubricas, ai.qtd_remuneracoes, ai.qtd_pagamentos, ai.warnings
        FROM arquivos_importados ai
        LEFT JOIN empresas e ON e.id = ai.company_id
        ORDER BY ai.imported_at DESC
        """,
        conn,
    )


def to_excel_bytes(sheets: Dict[str, pd.DataFrame]) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for name, df in sheets.items():
            safe_name = name[:31]
            df.to_excel(writer, index=False, sheet_name=safe_name)
            ws = writer.sheets[safe_name]
            for column_cells in ws.columns:
                max_len = 0
                letter = column_cells[0].column_letter
                for cell in column_cells:
                    value = "" if cell.value is None else str(cell.value)
                    max_len = max(max_len, min(len(value), 60))
                ws.column_dimensions[letter].width = max(10, max_len + 2)
    return output.getvalue()


def _br_money(value: float | int | None) -> str:
    value = 0.0 if value is None else float(value)
    return f"R$ {value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _format_cnpj(value: str | None) -> str:
    digits = "".join(ch for ch in str(value or "") if ch.isdigit())
    if len(digits) == 14:
        return f"{digits[:2]}.{digits[2:5]}.{digits[5:8]}/{digits[8:12]}-{digits[12:]}"
    return digits or "-"


def _month_label(value: str | int | float | None) -> str:
    text = "" if value is None else str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    if text == "13":
        return "13 - 13o salario"
    if text.isdigit():
        return f"Mes {int(text):02d}"
    return text or "-"


def to_perdcomp_simple_excel_bytes(monthly: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "PERDCOMP"

    red_fill = PatternFill("solid", fgColor="E21B3C")
    dark_fill = PatternFill("solid", fgColor="1F2937")
    light_fill = PatternFill("solid", fgColor="F3F4F6")
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(["Relatorio simplificado PER/DCOMP"])
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = red_fill
    ws.merge_cells("A1:D1")

    row_idx = 3
    if monthly.empty:
        ws.cell(row=row_idx, column=1, value="Nenhum credito mensal encontrado.")
    else:
        data = monthly.copy()
        data["credito_estimado"] = pd.to_numeric(data["credito_estimado"], errors="coerce").fillna(0.0)
        company_cols = ["razao_social", "cnpj_completo"]
        for (company_name, cnpj), company_df in data.groupby(company_cols, dropna=False):
            company_name = company_name or "Empresa sem nome"
            cnpj_label = _format_cnpj(cnpj)
            total_company = company_df["credito_estimado"].sum()

            ws.cell(row=row_idx, column=1, value="Empresa")
            ws.cell(row=row_idx, column=2, value=company_name)
            ws.cell(row=row_idx, column=3, value="CNPJ")
            ws.cell(row=row_idx, column=4, value=cnpj_label)
            for cell in ws[row_idx]:
                cell.font = Font(bold=True)
                cell.fill = light_fill
                cell.border = border
            row_idx += 1

            ws.cell(row=row_idx, column=1, value="Credito total")
            ws.cell(row=row_idx, column=2, value=total_company)
            ws.cell(row=row_idx, column=2).number_format = '"R$" #,##0.00'
            ws.cell(row=row_idx, column=1).font = Font(bold=True, color="E21B3C")
            ws.cell(row=row_idx, column=2).font = Font(bold=True, color="E21B3C")
            row_idx += 2

            for year in sorted(company_df["ano"].dropna().unique()):
                year_df = company_df[company_df["ano"] == year].copy()
                year_total = year_df["credito_estimado"].sum()
                ws.cell(row=row_idx, column=1, value=f"Ano apurado {year}")
                ws.cell(row=row_idx, column=2, value=year_total)
                ws.cell(row=row_idx, column=2).number_format = '"R$" #,##0.00'
                for cell in ws[row_idx]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = dark_fill
                    cell.border = border
                row_idx += 1

                ws.append(["Mes", "Periodo apurado", "Periodo pago", "Credito"])
                for cell in ws[row_idx]:
                    cell.font = Font(bold=True)
                    cell.fill = light_fill
                    cell.border = border
                row_idx += 1

                year_df["_mes_ordem"] = pd.to_numeric(year_df["mes"], errors="coerce").fillna(99)
                for _, item in year_df.sort_values(["_mes_ordem", "periodo_apurado"]).iterrows():
                    ws.cell(row=row_idx, column=1, value=_month_label(item.get("mes")))
                    ws.cell(row=row_idx, column=2, value=item.get("periodo_apurado"))
                    ws.cell(row=row_idx, column=3, value=item.get("periodos_pagamento") or "-")
                    ws.cell(row=row_idx, column=4, value=float(item.get("credito_estimado") or 0.0))
                    ws.cell(row=row_idx, column=4).number_format = '"R$" #,##0.00'
                    for cell in ws[row_idx]:
                        cell.border = border
                    row_idx += 1
                row_idx += 1
            row_idx += 1

    widths = [24, 18, 20, 18]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center")

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def to_perdcomp_simple_pdf_bytes(monthly: pd.DataFrame) -> bytes:
    output = BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=A4,
        rightMargin=1.4 * cm,
        leftMargin=1.4 * cm,
        topMargin=1.3 * cm,
        bottomMargin=1.3 * cm,
    )
    styles = getSampleStyleSheet()
    story: list[Any] = [
        Paragraph("Relatorio simplificado PER/DCOMP", styles["Title"]),
        Spacer(1, 0.25 * cm),
    ]

    if monthly.empty:
        story.append(Paragraph("Nenhum credito mensal encontrado.", styles["Normal"]))
    else:
        data = monthly.copy()
        data["credito_estimado"] = pd.to_numeric(data["credito_estimado"], errors="coerce").fillna(0.0)
        for (company_name, cnpj), company_df in data.groupby(["razao_social", "cnpj_completo"], dropna=False):
            company_name = company_name or "Empresa sem nome"
            cnpj_label = _format_cnpj(cnpj)
            total_company = company_df["credito_estimado"].sum()
            story.append(Paragraph(f"<b>Empresa:</b> {company_name}", styles["Heading3"]))
            story.append(Paragraph(f"<b>CNPJ:</b> {cnpj_label}", styles["Normal"]))
            story.append(Paragraph(f"<b>Credito total:</b> {_br_money(total_company)}", styles["Heading3"]))
            story.append(Spacer(1, 0.15 * cm))

            for year in sorted(company_df["ano"].dropna().unique()):
                year_df = company_df[company_df["ano"] == year].copy()
                year_total = year_df["credito_estimado"].sum()
                story.append(Paragraph(f"Ano apurado {year} - Credito {_br_money(year_total)}", styles["Heading4"]))
                table_rows = [["Mes", "Periodo apurado", "Periodo pago", "Credito"]]
                year_df["_mes_ordem"] = pd.to_numeric(year_df["mes"], errors="coerce").fillna(99)
                for _, item in year_df.sort_values(["_mes_ordem", "periodo_apurado"]).iterrows():
                    table_rows.append(
                        [
                            _month_label(item.get("mes")),
                            str(item.get("periodo_apurado") or "-"),
                            str(item.get("periodos_pagamento") or "-"),
                            _br_money(item.get("credito_estimado")),
                        ]
                    )
                table = Table(table_rows, colWidths=[3.2 * cm, 4.0 * cm, 4.0 * cm, 4.0 * cm], repeatRows=1)
                table.setStyle(
                    TableStyle(
                        [
                            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
                            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D1D5DB")),
                            ("ALIGN", (3, 1), (3, -1), "RIGHT"),
                            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
                        ]
                    )
                )
                story.append(table)
                story.append(Spacer(1, 0.35 * cm))
            story.append(Spacer(1, 0.2 * cm))

    doc.build(story)
    return output.getvalue()
